#!/usr/bin/env python3
from sysconfig import get_scheme_names
import tkinter as tk
from tkinter import *
import threading
import serial
import os
import time
from tkinter import messagebox
import json
import configparser
import openpyxl 
import RPi.GPIO as GPIO
import esptool
w=0
h=0

LARGEFONT =0
MEDIUMFONT =0

GPIO_Ras2Holtek_1=5	 	#pin 29 : Ras2Holtek-1
GPIO_Ras2Holtek_2=6	 	#pin 31 : Ras2Holtek-2
GPIO_H_ExtTrigger=19	#pin 35 : H-ExtTrigger
GPIO_H_check_out=26		#pin 37 : H-check-out
GPIO_ESP_EN=23			#pin 16 : GPIO_ESP_EN
GPIO_ESP_RST=24			#pin 18 : GPIO_ESP_RST


Button_list = []
Check_list = []
Serial_label=[]


Test_Tread=threading.Thread()
Test_Tread_Stop=0
Test_Mode_Auto=0
Test_index=-1



def xx(precent):
	return int(w*precent/100)
	
def yy(precent):
	return int(h*precent/100)

def hh(precent):
	return int(h*precent/1725)
	
def ww(precent):
	return int(w*precent/803)	
	

def elapsed():
	return int(round(time.time() * 1000))

def init_scale():
	global w,h,LARGEFONT,MEDIUMFONT
	tmp=tk.Tk()
	w = tmp.winfo_screenwidth()
	h = tmp.winfo_screenheight()
	tmp.destroy()
	LARGEFONT =("Verdana", int(w/24))
	MEDIUMFONT =("Verdana", int(w/150))




def Command_json(action,mode,timeout):
	  
	string="{action:\""+action+"\",mode:\""+mode+"\",timeout:\""+str(timeout)+"\"}\r\n"
	
	print("Frame_STM="+string)
	try:
		serial_stm = serial.Serial ("/dev/ttyS0", 115200,timeout=0)		
		serial_stm.write(string.encode('ascii'))
	except:
		messagebox.showinfo("Error!", "Serial Port Error!")
		return 
	
	current=elapsed()
	while elapsed()-current < timeout+100:
		time.sleep(0.1)
	try:
		received_data = serial_stm.read(1000)
	except:
		messagebox.showinfo("Error!", "Serial Port Error!")
		return 
	
	res=received_data.decode("utf-8")
	#res="{\"response\":\"ok\"}"
	print("return_STM="+res)

	try:
		j=json.loads(res)
		return j["response"]
	except:
		return ""
			


def receive_json(timeout):
	
	serial_stm = serial.Serial ("/dev/ttyS0", 115200,timeout=0)
	current=elapsed()
	while elapsed()-current < timeout+100:
		time.sleep(0.1)
	try:
		
		received_data = serial_stm.read(1000)
		res=received_data.decode("utf-8")
		print("return_STM="+res)
		try:
			j=json.loads(res)
			return j["response"]
		except:
			return ""
			
	except:
		messagebox.showinfo("Error!", "Serial Port Error!")
		return 

def send_json(action,mode,timeout):
	  
	string="{action:\""+action+"\",mode:\""+mode+"\",timeout:\""+str(timeout)+"\"}\r\n"
	
	print("Frame_STM="+string)
	try:
		serial_stm = serial.Serial ("/dev/ttyS0", 115200,timeout=0)		
		serial_stm.write(string.encode('ascii'))
	except:
		messagebox.showinfo("Error!", "Serial Port Error!")
		return 



def Command_json_esp(action,mode,timeout,phone):
	  
	string="{"
	if action:
		string=string+"\"action\":\""+action+"\""
	if mode:
		string=string+",\"mode\":\""+mode+"\""
	string=string+",\"timeout\":"+str(timeout)
	if phone:
		string=string+",\"phone\":\""+phone+"\""
	string=string+"}\r\n"
	
	print("Frame_ESP="+string)
	try:
		serial_esp = serial.Serial ("/dev/ttyUSB0", 115200,timeout=0)		
		serial_esp.write(string.encode('ascii'))
		current=elapsed()
		while elapsed()-current < timeout*1000+600:
			time.sleep(0.1)
		received_data = serial_esp.read(1000)
		res=received_data.decode("utf-8")
		
		
		#res="res=\n{\"response\": \"10000001\"}"
		try:
			sp=res.splitlines()
			print("return_ESP="+sp[1])

			j=json.loads(sp[1])
			return j["response"]
		except:
			return ""
			
	except:
		messagebox.showinfo("Error!", "Serial Port Error!")
		return 


def Test_Start(index):
	global Test_index,Test_Tread_Stop

	if index >1 and Serial_label[1]=="":
		#messagebox.showinfo("Error!", "You must do the SerialNumber test first!")
		#return
		Serial_label[1]="device1"
		Serial_label[0].config(text = "SN: "+Serial_label[1])
		

	Button_list[index].configure(bg = "orange",activebackground = "orange")	
	for button in Button_list:
		button["state"] = DISABLED
	for check in Check_list:
		check["state"] = DISABLED

	Button_list[18]["text"]="Stop"  
	Button_list[18]["state"]=NORMAL

	Test_index=index
	Test_Tread=threading.Thread(target=Test_list[index])
	Test_Tread_Stop=0
	Test_Tread.start()



def Test_End(result):
	global Test_index
	
	if result==1:   	#Success test
		Button_list[Test_index].configure(bg = "green",activebackground = "green")
	elif result==2:		#Fail test 
		Button_list[Test_index].configure(bg = "red",activebackground = "red")
	else:				#Cancelled 
		Button_list[Test_index].configure(bg = "azure2",activebackground = "azure2")

	if Test_Mode_Auto==0:
		for button in Button_list:
			button["state"] = NORMAL
		for check in Check_list:
			check["state"] = NORMAL
		Button_list[18]["text"]="Start"

	Test_index=-1


def Test_Start_Auto():
	global Test_index,Test_Mode_Auto

	for button in Button_list:
		button["bg"] = "azure2"
	
	Test_Mode_Auto=1

	for index in range(0,18):
		Test_Start(index)
		while Test_index >= 0:
			if Test_Tread_Stop:
				Test_Mode_Auto=0
				return
			pass
	Test_Mode_Auto=0
	for button in Button_list:
		button["state"] = NORMAL
	for check in Check_list:
		check["state"] = NORMAL
	Button_list[18]["text"]="Start"




def Start_Stop():
	global Test_index,Test_Tread_Stop
	if Test_index >= 0:
		Test_Tread_Stop=1
	else:
		Test_Tread=threading.Thread(target=Test_Start_Auto)
		Test_Tread.start()
		

def get_sn():
	path="/boot/Tester/serial.xlsx"
	try:
		wb_obj = openpyxl.load_workbook(path)	
		sheet_obj = wb_obj.active
	except:		
		messagebox.showinfo("Error!", "Cannot open excel file !")
		return False

	row = sheet_obj.max_row

	n=0
	for i in range(2,row+1):
		used=sheet_obj.cell(row=i, column=2).value
		if used==0:
			n=i
			break

	if n==0:
		messagebox.showinfo("Error!", "All of serials in excel file is used!")
		return False
	
	ret=[]
	ret.append(str(sheet_obj.cell(row=n, column=1).value))
	ret.append(n)

	return ret


def set_sn_used(ind):
	path="/boot/Tester/serial.xlsx"
	try:
		wb_obj = openpyxl.load_workbook(path)	
		sheet_obj = wb_obj.active
	except:		
		messagebox.showinfo("Error!", "Cannot open excel file !")
		return False

	sheet_obj.cell(row=ind, column=2).value=1
	# ind='B'+str(n)
	# sheet_obj[ind].value=1
	wb_obj.save(path)


def Program_t():
	global Test_Tread_Stop		
	config = configparser.ConfigParser() 
	config2 = configparser.ConfigParser() 
	config.read('config.ini')
	config2.read('results.ini')
	try:
		config2.add_section("program")
	except:
		pass

	values=[]
	results=[]


	ret=1
	timeout=config.get("timeouts","time_0")
	res=Command_json("power swith","1",int(timeout))	
	time_program=int(config.get("timeouts","time_1"))
	if res== "ok":
		GPIO.setmode(GPIO.BCM)
		GPIO.setup(GPIO_Ras2Holtek_1, GPIO.OUT) 
		GPIO.setup(GPIO_Ras2Holtek_2, GPIO.OUT)  
		GPIO.setup(GPIO_H_ExtTrigger, GPIO.OUT)  
		GPIO.setup(GPIO_H_check_out, GPIO.IN, pull_up_down=GPIO.PUD_UP)  

		GPIO.output(GPIO_H_ExtTrigger, GPIO.LOW)
		GPIO.output(GPIO_Ras2Holtek_1, GPIO.HIGH)
		GPIO.output(GPIO_Ras2Holtek_2, GPIO.LOW)

		GPIO.output(GPIO_H_ExtTrigger, GPIO.HIGH)
		current=elapsed()
		res="Timeout"
		while elapsed()-current < time_program:
			time.sleep(0.1)
			if Test_Tread_Stop:
				Test_End(0)
				return
			if GPIO.input(GPIO_H_check_out):
				res="OK"
				break
		values.append(res)
		if res=="OK":
			results.append(1)
		else:
			results.append(0)
			ret=2

		GPIO.output(GPIO_H_ExtTrigger, GPIO.LOW)
		GPIO.output(GPIO_Ras2Holtek_1, GPIO.LOW)
		time.sleep(0.5)
		GPIO.output(GPIO_Ras2Holtek_2, GPIO.HIGH)
		GPIO.output(GPIO_H_ExtTrigger, GPIO.HIGH)
		current=elapsed()
		res="Timeout"
		while elapsed()-current < time_program:
			time.sleep(0.1)
			if Test_Tread_Stop:
				Test_End(0)
				return
			if GPIO.input(GPIO_H_check_out):
				res="OK"
				break
		values.append(res)
		if res=="OK":
			results.append(1)
		else:
			results.append(0)
			ret=2
		GPIO.output(GPIO_H_ExtTrigger, GPIO.LOW)
		GPIO.output(GPIO_Ras2Holtek_1, GPIO.LOW)
		GPIO.output(GPIO_Ras2Holtek_2, GPIO.LOW)

	else:
		values=["STM power switch Err"]*2
		results=[0]*2
		ret=2
	if Test_Tread_Stop:
		Test_End(0)
		return

	res=Command_json("power swith","0",int(timeout))
	if res== "ok":
		GPIO.setup(GPIO_ESP_EN, GPIO.OUT)  
		GPIO.output(GPIO_ESP_EN, GPIO.HIGH)
		GPIO.setup(GPIO_ESP_RST, GPIO.OUT)  
		GPIO.output(GPIO_ESP_RST, GPIO.HIGH)

		GPIO.output(GPIO_ESP_EN, GPIO.LOW)

		GPIO.output(GPIO_ESP_RST, GPIO.LOW)
		time.sleep(3)
		GPIO.output(GPIO_ESP_RST, GPIO.HIGH)

		path='/boot/Tester/esp.bin'
		command = ['--baud', '460800','--chip' ,'esp32', 'write_flash', '-z' ,'0x0000' ,path]
		result_esp=""
		try:
			ret=esptool.main(command)
			result_esp="OK"
		except Exception as e:
			result_esp="ESP flash Err"

		GPIO.output(GPIO_ESP_EN, GPIO.HIGH)

		GPIO.output(GPIO_ESP_RST, GPIO.LOW)
		time.sleep(1)
		GPIO.output(GPIO_ESP_RST, GPIO.HIGH)

		values.append(result_esp)
		if result_esp=="OK":
			results.append(1)
		else:
			results.append(0)
			ret=2

		values.append("OK")
		results.append(1)
		

	else:
		values.append("STM power switch Err")
		values.append("STM power switch Err")
		results.append(0)
		results.append(0)
		ret=2



	config2.set("program","program_value",json.dumps(values))
	config2.set("program","program_result",json.dumps(results))

	with open('results.ini', 'w') as configfile:
		config2.write(configfile)

	Test_End(ret)



def SerialNumber_t():
	global Test_Tread_Stop

	ret=1
	res=Command_json_esp("serial","read",0,False)
	if res:
		Serial_label[1]=res
		Serial_label[0].config(text = "SN: "+Serial_label[1])
	else:
		sn=get_sn()	
		if sn:
			res=Command_json_esp("serial","set",0,sn[0])
			if res=="OK":
				Serial_label[1]=sn[0]
				Serial_label[0].config(text = "SN: "+Serial_label[1])
				set_sn_used(sn[1])

				config = configparser.ConfigParser() 
				config.read('results.ini')
				try:
					config.add_section(Serial_label[1])
				except:
					pass
				values=config.get("program","program_value")
				results=config.get("program","program_result")
				config.set(Serial_label[1],"program_value",json.dumps(values))
				config.set(Serial_label[1],"program_result",json.dumps(results))
				with open('results.ini', 'w') as configfile:
					config.write(configfile)

			else:
				ret=2
		else:
			ret=2

	Test_End(ret)
		


def Supply_t():
	global Test_Tread_Stop
	config = configparser.ConfigParser() 
	config2 = configparser.ConfigParser() 
	config.read('config.ini')
	config2.read('results.ini')
	try:
		config2.add_section(Serial_label[1])
	except:
		pass

	time=config.get("timeouts","time_2")
	res=Command_json("read power","all",int(time))	
	values=[]
	results=[]

	try:
		values=json.loads(res)
		config2.set(Serial_label[1],"supply_value",json.dumps(values))
		re=1	
		for i in range(0,8):
			mn=config.get("ranges","min_"+str(i))
			mx=config.get("ranges","max_"+str(i))
			if(values[i]>int(mx) or values[i]<int(mn)):
				results.append(0)
				re=2
			else:
				results.append(1)
		config2.set(Serial_label[1],"supply_result",json.dumps(results))
	except:
		config2.set(Serial_label[1],"supply_value",json.dumps(['STM Err'] * 8))
		config2.set(Serial_label[1],"supply_result",json.dumps([0] * 8))
		re=2

	with open('results.ini', 'w') as configfile:
		config2.write(configfile)
	
	Test_End(re)



def Outputs_t():
	global Test_Tread_Stop
	config = configparser.ConfigParser() 
	config2 = configparser.ConfigParser() 
	config.read('config.ini')
	config2.read('results.ini')
	try:
		config2.add_section(Serial_label[1])
	except:
		pass

	time=config.get("timeouts","time_8")
	
	ret=Command_json_esp("output","High",0,False)
	if ret!="OK":
		config2.set(Serial_label[1],"output_value",json.dumps(['ESP Err'] * 4))
		config2.set(Serial_label[1],"output_result",json.dumps([0] * 4))
		res=2
	else:	
		res=Command_json("output","all",int(time))	
		values=[]
		results=[]
		try:
			values=json.loads(res)
			config2.set(Serial_label[1],"output_value",json.dumps(values))
			res=1
			for i in range(0,4):
				mn=config.get("ranges","min_"+str(i+14))
				mx=config.get("ranges","max_"+str(i+14))
				if(values[i]>int(mx) or values[i]<int(mn)):
					results.append(0)
					res=2
				else:
					results.append(1)
			config2.set(Serial_label[1],"output_result",json.dumps(results))
		except:
			config2.set(Serial_label[1],"output_value",json.dumps(['STM Err'] * 4))
			config2.set(Serial_label[1],"output_result",json.dumps([0] * 4))
			res=2
	with open('results.ini', 'w') as configfile:
		config2.write(configfile)
	ret=Command_json_esp("output","Low",1,False)
	Test_End(res)



def InputsZones_t():
	global Test_Tread_Stop
	config = configparser.ConfigParser() 
	config2 = configparser.ConfigParser() 
	config.read('config.ini')
	config2.read('results.ini')
	try:
		config2.add_section(Serial_label[1])
	except:
		pass

	time=config.get("timeouts","time_5")
	send_json("zone","all",int(time))	
	res_esp=Command_json_esp("zone","all",0,False)
	res_stm=receive_json(int(time))

	values_stm=[]
	values_esp=[]
	results_stm=[]
	results_esp=[]
	results=[]
	
	try:
		values_esp=json.loads(res_esp)
		config2.set(Serial_label[1],"zone_value_esp",json.dumps(values_esp))
		for i in range(0,5):
			if values_esp[i] == 0:
				results_esp.append(0)
			else:
				results_esp.append(1)
		config2.set(Serial_label[1],"zone_result_esp",json.dumps(results_esp))
	except:
		config2.set(Serial_label[1],"zone_value_esp",json.dumps(['ESP Err']* 5))
		config2.set(Serial_label[1],"zone_result_esp",json.dumps([0] * 5))
		values_esp=[0] * 5
		
	
	try:
		values_stm=json.loads(res_stm)
		config2.set(Serial_label[1],"zone_value_stm",json.dumps(values_stm))
		for i in range(0,5):
			mn=config.get("ranges","min_"+str(i+9))
			mx=config.get("ranges","max_"+str(i+9))
			if(values_stm[i]>int(mx) or values_stm[i]<int(mn)):
				results_stm.append(0)
			else:
				results_stm.append(1)
		config2.set(Serial_label[1],"zone_result_stm",json.dumps(results_stm))
	except:
		config2.set(Serial_label[1],"zone_value_stm",json.dumps(['STM Err'] * 5))
		config2.set(Serial_label[1],"zone_result_stm",json.dumps([0] * 5))
		results_stm=[0] * 5

	res=0
	for i in range(0,5):
		if results_stm[i]==1 and values_esp[i]==1:
			results.append(1)
		else:
			results.append(0)
			res=2
	config2.set(Serial_label[1],"zone_result",json.dumps(results))
	

	with open('results.ini', 'w') as configfile:
		config2.write(configfile)
	
	Test_End(res)
	


def Audio_t():
	global Test_Tread_Stop
	config = configparser.ConfigParser() 
	config2 = configparser.ConfigParser() 
	config.read('config.ini')
	config2.read('results.ini')
	try:
		config2.add_section(Serial_label[1])
	except:
		pass


	modes1=["sp1 f1","sp1 f2","sp1 f3"]
	modes2=["sp2 f1","sp2 f2","sp2 f3"]
	time=config.get("timeouts","time_10")

	min_holtek1=config.get("ranges","min_19")
	max_holtek1=config.get("ranges","max_19")
	min_holtek2=config.get("ranges","min_20")
	max_holtek2=config.get("ranges","max_20")

	min_signal1=config.get("ranges","min_21")
	max_signal1=config.get("ranges","max_21")
	min_signal2=config.get("ranges","min_22")
	max_signal2=config.get("ranges","max_22")

	min_present1=config.get("ranges","min_23")
	max_present1=config.get("ranges","max_23")
	min_present2=config.get("ranges","min_24")
	max_present2=config.get("ranges","max_24")

	min_siren=config.get("ranges","min_18")
	max_siren=config.get("ranges","max_18")



	result_user=[]
	results=[]

	values_esp=[]
	values_stm_holtek=[]
	values_stm_signal=[]
	values_stm_present=[]
	values_stm_speaker=[]
	values_stm_siren=[]

	results_esp=[]
	results_stm_holtek=[]
	results_stm_signal=[]
	results_stm_present=[]
	results_stm_speaker=[]
	results_stm_siren=[]


	Command_json_esp("speaker","sp1 f0",0,False)
	Command_json_esp("speaker","sp2 f0",0,False)


	for mod in modes1:
		if Test_Tread_Stop:
			Test_End(0)
			return
		res_esp=Command_json_esp("speaker",mod,0,False)		
		res_stm_signal=Command_json("sound","speaker signal",int(time))
		res_stm_holtek=Command_json("sound","holtek output",int(time))
		res_stm_present=Command_json("sound","speaker present",int(time))
		res_stm_speaker=Command_json("sound","speaker output",int(time))
		mss="Do you Hear "+mod+"?"	
		res_user=messagebox.askyesno("Audio test", mss)
		if res_user==True:
			result_user.append(1)
		else:
			result_user.append(0)
		
		if res_esp == "OK":
			results_esp.append(1)
			values_esp.append(res_esp)
		else:
			results_esp.append(0)
			values_esp.append("ESP Err")

		if res_stm_speaker == "ok":
			results_stm_speaker.append(1)
			values_stm_speaker.append(res_stm_speaker)
		else:
			results_stm_speaker.append(0)
			values_stm_speaker.append("STM Err")

		try:
			values_stm=json.loads(res_stm_holtek)
			if(values_stm[0]>int(max_holtek1) or values_stm[0]<int(min_holtek1)):
				results_stm_holtek.append(0)
			else:
				results_stm_holtek.append(1)
			values_stm_holtek.append(values_stm[0])
		except:
			results_stm_holtek.append(0)
			values_stm_holtek.append("STM Err")

		try:
			values_stm=json.loads(res_stm_signal)
			if(values_stm[0]>int(max_signal1) or values_stm[0]<int(min_signal1)):
				results_stm_signal.append(0)
			else:
				results_stm_signal.append(1)
			values_stm_signal.append(values_stm[0])
		except:
			results_stm_signal.append(0)
			values_stm_signal.append("STM Err")

		try:
			values_stm=json.loads(res_stm_present)
			if(values_stm[0]>int(max_present1) or values_stm[0]<int(min_present1)):
				results_stm_present.append(0)
			else:
				results_stm_present.append(1)
			values_stm_present.append(values_stm[0])
		except:
			results_stm_present.append(0)
			values_stm_present.append("STM Err")


	Command_json_esp("speaker","sp1 f0",0,False)
	Command_json_esp("speaker","sp2 f0",0,False)


	for mod in modes2:
		if Test_Tread_Stop:
			Test_End(0)
			return
		res_esp=Command_json_esp("speaker",mod,0,False)		
		res_stm_signal=Command_json("sound","speaker signal",int(time))
		res_stm_holtek=Command_json("sound","holtek output",int(time))
		res_stm_present=Command_json("sound","speaker present",int(time))
		res_stm_speaker=Command_json("sound","speaker output",int(time))
		mss="Do you Hear "+mod+"?"	
		res_user=messagebox.askyesno("Audio test", mss)
		if res_user==True:
			result_user.append(1)
		else:
			result_user.append(0)
		
		if res_esp == "OK":
			results_esp.append(1)
			values_esp.append(res_esp)
		else:
			results_esp.append(0)
			values_esp.append("ESP Err")

		if res_stm_speaker == "ok":
			results_stm_speaker.append(1)
			values_stm_speaker.append(res_stm_speaker)
		else:
			results_stm_speaker.append(0)
			values_stm_speaker.append("STM Err")

		try:
			values_stm=json.loads(res_stm_holtek)
			if(values_stm[1]>int(max_holtek2) or values_stm[1]<int(min_holtek2)):
				results_stm_holtek.append(0)
			else:
				results_stm_holtek.append(1)
			values_stm_holtek.append(values_stm[1])
		except:
			results_stm_holtek.append(0)
			values_stm_holtek.append("STM Err")

		try:
			values_stm=json.loads(res_stm_signal)
			if(values_stm[1]>int(max_signal2) or values_stm[1]<int(min_signal2)):
				results_stm_signal.append(0)
			else:
				results_stm_signal.append(1)
			values_stm_signal.append(values_stm[1])
		except:
			results_stm_signal.append(0)
			values_stm_signal.append("STM Err")

		try:
			values_stm=json.loads(res_stm_present)
			if(values_stm[1]>int(max_present2) or values_stm[1]<int(min_present2)):
				results_stm_present.append(0)
			else:
				results_stm_present.append(1)
			values_stm_present.append(values_stm[1])
		except:
			results_stm_present.append(0)
			values_stm_present.append("STM Err")


	Command_json_esp("speaker","sp1 f0",0,False)
	Command_json_esp("speaker","sp2 f0",0,False)


	if Test_Tread_Stop:
			Test_End(0)
			return

	time=config.get("timeouts","time_9")  #Read timeout for siren test
	res_esp=Command_json_esp("siren","play",0,False)		
	res_stm_siren=Command_json("siren","test",int(time))
	if res_esp == "OK":
		results_esp.append(1)
		values_esp.append(res_esp)
	else:
		results_esp.append(0)
		values_esp.append("ESP Err")
	try:
		values_stm=int(res_stm_siren)
		if(values_stm>int(max_siren) or values_stm<int(min_siren)):
			results_stm_siren.append(0)
		else:
			results_stm_siren.append(1)
		values_stm_siren.append(values_stm)
	except:
		results_stm_siren.append(0)
		values_stm_siren.append("STM Err")

	res_user=messagebox.askyesno("Audio test", "Do you Hear siren sound?")
	if res_user==True:
		result_user.append(1)
	else:
		result_user.append(0)

	res=0
	for i in range(0,6):
		if results_stm_present[i]==1 and results_stm_signal[i]==1 and results_stm_holtek[i]==1 and results_stm_speaker[i]==1 and results_esp[i]==1 and result_user[i]==1:
			results.append(1)
		else:
			results.append(0)
			res=2

	if results_stm_siren[0]==1 and results_esp[6]==1 and result_user[6]==1:
		results.append(1)
	else:
		results.append(0)
		res=2

	

	config2.set(Serial_label[1],"Audio_result",json.dumps(results))
	config2.set(Serial_label[1],"Audio_result_present",json.dumps(results_stm_present))
	config2.set(Serial_label[1],"Audio_result_signal",json.dumps(results_stm_signal))
	config2.set(Serial_label[1],"Audio_result_holtek",json.dumps(results_stm_holtek))
	config2.set(Serial_label[1],"Audio_result_speaker",json.dumps(results_stm_speaker))
	config2.set(Serial_label[1],"Audio_result_siren",json.dumps(results_stm_siren))
	config2.set(Serial_label[1],"Audio_result_esp",json.dumps(results_esp))
	config2.set(Serial_label[1],"Audio_result_user",json.dumps(result_user))

	config2.set(Serial_label[1],"Audio_values_present",json.dumps(values_stm_present))
	config2.set(Serial_label[1],"Audio_values_signal",json.dumps(values_stm_signal))
	config2.set(Serial_label[1],"Audio_values_holtek",json.dumps(values_stm_holtek))
	config2.set(Serial_label[1],"Audio_values_speaker",json.dumps(values_stm_speaker))
	config2.set(Serial_label[1],"Audio_values_siren",json.dumps(values_stm_siren))
	config2.set(Serial_label[1],"Audio_values_esp",json.dumps(values_esp))

	
	with open('results.ini', 'w') as configfile:
		config2.write(configfile)
	
	Test_End(res)

	
	
def Dummy_t():
	global Test_Tread_Stop
	config = configparser.ConfigParser() 
	config2 = configparser.ConfigParser() 
	config.read('config.ini')
	config2.read('results.ini')
	try:
		config2.add_section(Serial_label[1])
	except:
		pass

	time=config.get("timeouts","time_7")
	min=config.get("ranges","min_8")
	max=config.get("ranges","max_8")
	
	values_stm1=[]
	results_stm1=[]
	values_stm2=[]
	results_stm2=[]


	re=1
	values_stm1=Command_json("dummy load","1",int(time))	

	try:
		values_stm1=json.loads(values_stm1)
		for val in values_stm1:
			if(val>int(max) or val<int(min)):
				results_stm1.append(0)
				re=2
			else:
				results_stm1.append(1)
	except:
		results_stm1=[0] * 2
		values_stm1=['STM Err'] * 2
		re=2

	if Test_Tread_Stop:
		Test_End(0)
		return

	values_stm2=Command_json("dummy load","2",int(time))	

	try:
		values_stm2=json.loads(values_stm2)
		for val in values_stm2:
			if(val>int(max) or val<int(min)):
				results_stm2.append(0)
				re=2
			else:
				results_stm2.append(1)
	except:
		results_stm2=[0] * 2
		values_stm2=['STM Err'] * 2
		re=2


	config2.set(Serial_label[1],"Dummy_value_stm_1",json.dumps(values_stm1))	
	config2.set(Serial_label[1],"Dummy_result_stm_1",json.dumps(results_stm1))

	config2.set(Serial_label[1],"Dummy_value_stm_2",json.dumps(values_stm2))	
	config2.set(Serial_label[1],"Dummy_result_stm_2",json.dumps(results_stm2))

	with open('results.ini', 'w') as configfile:
		config2.write(configfile)
	
	Test_End(re)



def ShortCircuits_t():
	global Test_Tread_Stop
	config = configparser.ConfigParser() 
	config2 = configparser.ConfigParser() 
	config.read('config.ini')
	config2.read('results.ini')
	try:
		config2.add_section(Serial_label[1])
	except:
		pass

	time=config.get("timeouts","time_3")
	
	values_stm=[]
	results_stm=[]
	values_esp=[]
	results_esp=[]

	modes=["speaker1","speaker2","external voltage","siren","battery"]
	response=["A","B","C","D","E"]

	re=1
	cnt=0
	for mod in modes:
		if Test_Tread_Stop:
			Test_End(0)
			return
		res_stm=Command_json("short circuit",mod,int(time))			
		if res_stm=="ok":
			values_stm.append(res_stm)
			results_stm.append(1)
		elif res_stm=="":
			values_stm.append('STM Err')
			results_stm.append(0)
			re=2
		else:
			values_stm.append(res_stm)
			results_stm.append(0)
			re=2

		res_esp=Command_json_esp("panel_control",False,0,False)
		if res_esp==response[cnt]:
			values_esp.append(res_stm)
			results_esp.append(1)
		elif res_esp=="":
			values_esp.append('ESP Err')
			results_esp.append(0)
			re=2
		else:
			values_esp.append(res_stm)
			results_esp.append(0)
			re=2


		
			

	config2.set(Serial_label[1],"shortcircuit_value_stm",json.dumps(values_stm))	
	config2.set(Serial_label[1],"shortcircuit_result_stm",json.dumps(results_stm))
	config2.set(Serial_label[1],"shortcircuit_value_esp",json.dumps(values_esp))	
	config2.set(Serial_label[1],"shortcircuit_result_esp",json.dumps(results_esp))

	with open('results.ini', 'w') as configfile:
		config2.write(configfile)
	
	Test_End(re)

	

def Battery_t():
	global Test_Tread_Stop
	config = configparser.ConfigParser() 
	config2 = configparser.ConfigParser() 
	config.read('config.ini')
	config2.read('results.ini')
	try:
		config2.add_section(Serial_label[1])
	except:
		pass

	time=config.get("timeouts","time_2")
	min=config.get("ranges","min_7")
	max=config.get("ranges","max_7")
	
	values_stm=[]
	results_stm=[]
	values_esp=[]
	results_esp=[]

	re=1


	values_esp=Command_json_esp("power",False,0,False)

	try:
		#values_esp=json.loads(values_esp)		
		if values_esp[0] ==1:
			results_esp.append(1)
		else:
			results_esp.append(0)
			re=2
		if values_esp[1] ==1:
			results_esp.append(1)
		else:
			results_esp.append(0)
			re=2

	except:
		results_esp=[0] * 2
		values_esp=['ESP Err'] * 3

	if Test_Tread_Stop:
		Test_End(0)
		return

	

	values_stm=Command_json("read power","battery",int(time))	

	try:
		values_stm=json.loads(values_stm)
		for val in values_stm:
			if(val>int(max) or val<int(min)):
				results_stm.append(0)
				re=2
			else:
				results_stm.append(1)

	except:
		results_stm=[0] * 2
		values_stm=['STM Err'] * 2
		re=2


	config2.set(Serial_label[1],"battery_value_stm",json.dumps(values_stm))	
	config2.set(Serial_label[1],"battery_result_stm",json.dumps(results_stm))
	config2.set(Serial_label[1],"battery_value_esp",json.dumps(values_esp))	
	config2.set(Serial_label[1],"battery_result_esp",json.dumps(results_esp))

	with open('results.ini', 'w') as configfile:
		config2.write(configfile)
	
	Test_End(re)



def GSM_t():
	global Test_Tread_Stop
	config = configparser.ConfigParser() 
	config2 = configparser.ConfigParser() 
	config.read('config.ini')
	config2.read('results.ini')
	try:
		config2.add_section(Serial_label[1])
	except:
		pass

	time=config.get("timeouts","time_4")
	phone=config.get("ranges","phone")

	gsm_values_esp=[]
	gsm_results_esp=[]
	gsm_values_stm=[]
	gsm_results_stm=[]
	gsm_results_user=[]

	res_esp=Command_json_esp("simcard","initial",0,False)
	res=1 
	if res_esp=="OK":
		gsm_values_esp.append(res_esp)
		gsm_results_esp.append(1)		
	else:	
		gsm_values_esp.append("ESP Err")
		gsm_results_esp.append(0)
		res=2

	if Test_Tread_Stop:
		Test_End(0)
		return
	

	res_esp=Command_json_esp("simcard","get time",0,False)
	if res_esp=="OK":
		gsm_values_esp.append(res_esp)
		gsm_results_esp.append(1)		
	else:	
		gsm_values_esp.append("ESP Err")
		gsm_results_esp.append(0)
		res=2
	
	if Test_Tread_Stop:
		Test_End(0)
		return
	

	res_esp=Command_json_esp("simcard","sms test",0,phone)
	if res_esp=="OK":
		gsm_values_esp.append(res_esp)
		gsm_results_esp.append(1)		
	else:	
		gsm_values_esp.append("ESP Err")
		gsm_results_esp.append(0)
		res=2

	res_user=messagebox.askyesno("GSM test", "Did you get the sms?")
	

	if res_user==True:
		gsm_results_user.append(1)
	else:
		gsm_results_user.append(0)
		res=2

		
	if Test_Tread_Stop:
		Test_End(0)
		return
	

	res_stm=Command_json("microphone","test",int(time))
	
	if res_stm=="ok":
		gsm_values_stm.append(res_stm)
		gsm_results_stm.append(1)		
	else:	
		gsm_values_stm.append("STM Err")
		gsm_results_stm.append(0)
		res=2

	res_esp=Command_json_esp("simcard","call test",3,phone)
	if res_esp=="OK":
		gsm_values_esp.append(res_esp)
		gsm_results_esp.append(1)		
	else:	
		gsm_values_esp.append("ESP Err")
		gsm_results_esp.append(0)
		res=2

	res_user=messagebox.askyesno("GSM test", "Did you get the call?")
	if res_user==True:
		gsm_results_user.append(1)
	else:
		gsm_results_user.append(0)
		res=2


	config2.set(Serial_label[1],"gsm_values_esp",json.dumps(gsm_values_esp))	
	config2.set(Serial_label[1],"gsm_results_esp",json.dumps(gsm_results_esp))
	config2.set(Serial_label[1],"gsm_values_stm",json.dumps(gsm_values_stm))	
	config2.set(Serial_label[1],"gsm_results_stm",json.dumps(gsm_results_stm))
	config2.set(Serial_label[1],"gsm_results_user",json.dumps(gsm_results_user))

	with open('results.ini', 'w') as configfile:
		config2.write(configfile)

	Test_End(res)


def operator_t():
	global Test_Tread_Stop
	config = configparser.ConfigParser() 
	config2 = configparser.ConfigParser() 
	config.read('config.ini')
	config2.read('results.ini')
	try:
		config2.add_section(Serial_label[1])
	except:
		pass

	time=config.get("timeouts","time_6")

	operator_values_esp=[]
	operator_results_esp=[]
	operator_values_stm=[]
	operator_results_stm=[]
	operator_results_user=[]

	res_esp=Command_json_esp("panel","led on",0,False)
	res_user=messagebox.askyesno("Panel test", "All the panel leds are on?")

	res=1 
	if res_esp=="OK":
		operator_values_esp.append(res_esp)
		operator_results_esp.append(1)
		
	else:	
		operator_values_esp.append("ESP Err")
		operator_results_esp.append(0)
		res=2
		
	if res_user==True:
		operator_results_user.append(1)
	else:
		operator_results_user.append(0)
		res=2


	if Test_Tread_Stop:
		Test_End(0)
		return
	

	res_esp=Command_json_esp("panel","led off",0,False)
	res_user=messagebox.askyesno("Panel test", "All the panel leds are off?")

	res=1 
	if res_esp=="OK":
		operator_values_esp.append(res_esp)
		operator_results_esp.append(1)
		
	else:	
		operator_values_esp.append("ESP Err")
		operator_results_esp.append(0)
		res=2
		
	if res_user==True:
		operator_results_user.append(1)
	else:
		operator_results_user.append(0)
		res=2


	if Test_Tread_Stop:
			Test_End(0)
			return
	

	messagebox.showinfo("Starting 7segment test","In this test, you must see 0-F on segments!")


	modes=["0","1","2","3","4","5","6","7","8","9","a","b","c","d","e","f"]

	res_esp="OK"
	for mode in modes:
		res=Command_json_esp("segment",mode,0,False)
		if res!= "OK":
			res_esp=""
		if Test_Tread_Stop:
			Test_End(0)
			return

	
	res_user=messagebox.askyesno("Panel test", "All the segment modes was correct?")

	if res_esp=="OK":
		operator_values_esp.append(res_esp)
		operator_results_esp.append(1)
		
	else:	
		operator_values_esp.append("ESP Err")
		operator_results_esp.append(0)
		res=2
		
	if res_user==True:
		operator_results_user.append(1)
	else:
		operator_results_user.append(0)
		res=2


	res_esp=Command_json_esp("key","test",int(int(time)/1000)+1,False)
	res_stm=Command_json("key","test",time)

	res_user=messagebox.askyesno("Panel test", "Do you heard the Buzzer?")
	

	if res_esp=="OK":
		operator_values_esp.append(res_esp)
		operator_results_esp.append(1)
		
	else:	
		operator_values_esp.append("ESP Err")
		operator_results_esp.append(0)
		res=2

	if res_stm=="ok":
		operator_values_stm.append(res_esp)
		operator_results_stm.append(1)
		
	else:	
		operator_values_stm.append("STM Err")
		operator_results_stm.append(0)
		res=2
		
	if res_user==True:
		operator_results_user.append(1)
	else:
		operator_results_user.append(0)
		res=2


	config2.set(Serial_label[1],"operator_value_esp",json.dumps(operator_values_esp))	
	config2.set(Serial_label[1],"operator_results_esp",json.dumps(operator_results_esp))
	config2.set(Serial_label[1],"operator_values_stm",json.dumps(operator_values_stm))	
	config2.set(Serial_label[1],"operator_results_stm",json.dumps(operator_results_stm))
	config2.set(Serial_label[1],"operator_results_user",json.dumps(operator_results_user))

	with open('results.ini', 'w') as configfile:
		config2.write(configfile)

	Test_End(res)


def Function1_t():
	Test_End(0)


def Function2_t():
	Test_End(0)


def Function3_t():
	Test_End(0)


def Temperature_t():
	Test_End(0)


def Remote_t():
	Test_End(0)


def SoundGSM2_t():
	Test_End(0)


def FactorySet_t():
	global Test_Tread_Stop


	config2 = configparser.ConfigParser() 
	config2.read('results.ini')
	try:
		config2.add_section(Serial_label[1])
	except:
		pass


	factory_values_esp=[]
	factory_results_esp=[]


	res_esp=Command_json_esp("setDevice","time read",0,False)
	ret=1 
	if res_esp !="":
		factory_values_esp.append(res_esp)
		factory_results_esp.append(1)		
	else:	
		factory_values_esp.append("ESP Err")
		factory_results_esp.append(0)
		res=2
	
	res_esp=Command_json_esp("setDevice","save serial",0,False)
	if res_esp =="OK":
		factory_values_esp.append(res_esp)
		factory_results_esp.append(1)		
	else:	
		factory_values_esp.append("ESP Err")
		factory_results_esp.append(0)
		ret=2

	
	res_esp=Command_json_esp("setDevice","file set",0,False)
	if res_esp =="OK":
		factory_values_esp.append(res_esp)
		factory_results_esp.append(1)		
	else:	
		factory_values_esp.append("ESP Err")
		factory_results_esp.append(0)
		ret=2

	config2.set(Serial_label[1],"factory_value",json.dumps(factory_values_esp))
	config2.set(Serial_label[1],"factory_result",json.dumps(factory_results_esp))

	with open('results.ini', 'w') as configfile:
		config2.write(configfile)

	Test_End(ret)


Test_list=[Program_t,SerialNumber_t,Supply_t,Outputs_t,InputsZones_t,Audio_t
,Dummy_t,ShortCircuits_t,Battery_t,GSM_t,operator_t,Function1_t
,Function2_t,Function3_t,Temperature_t,Remote_t,SoundGSM2_t,FactorySet_t]


